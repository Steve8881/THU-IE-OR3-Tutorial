"""
Reading from and writing to an Excel file using openpyxl.
"""

from openpyxl import load_workbook

import constant


def readDataOpenpyxl() -> tuple[dict, dict, dict]:
    """
    Given an Excel file, read production planning data from Excel file using openpyxl.

    Returns
    -------
    productProfits : dict
        The profit per batch for each product
        - Keys: product names ("Doors", "Windows")
        - Values: profit per batch
    plantProductHours : nested dict
        The hours required to produce one batch of each product at each plant
        - Keys: plant names ("Plant 1", "Plant 2", "Plant 3")
        - Values: dict
            - Keys: product names ("Doors" and "Windows")
            - Values: hours required
    plantAvailableHours : dict
        The available hours at each plant
        - Keys: plant names ("Plant 1", "Plant 2", "Plant 3")
        - Values: available hours
    """

    # Load a workbook from DATA_PATH
    inputBook = load_workbook(constant.DATA_PATH)

    # Find the sheet in which you want to write the solution
    inputSheet = inputBook[constant.SHEET_NAME]

    # Read data from the inputSheet and create dictionaries
    productProfits = {}
    for productName in constant.PRODUCT_NAMES:
        profitColConst = getattr(constant, f"INPUT_PROFIT_{productName.upper()}_COL")
        profitValue = inputSheet.cell(constant.INPUT_PROFIT_ROW, profitColConst).value
        productProfits[productName] = profitValue

    plantAvailableHours = {}
    for i, plantName in enumerate(constant.PLANT_NAMES):
        rowIndex = constant.INPUT_HOURS_AVAILABLE_START_ROW + i
        colIndex = constant.INPUT_HOURS_AVAILABLE_COL
        hours = inputSheet.cell(rowIndex, colIndex).value
        plantAvailableHours[plantName] = hours

    plantProductHours = {}
    for i, plantName in enumerate(constant.PLANT_NAMES):
        rowIndex = constant.INPUT_HOURS_START_ROW + i
        hoursPerProduct = {}
        for productName in constant.PRODUCT_NAMES:
            hourColConst = getattr(constant, f"INPUT_HOURS_{productName.upper()}_COL")
            hours = inputSheet.cell(rowIndex, hourColConst).value
            hoursPerProduct[productName] = hours
        plantProductHours[plantName] = hoursPerProduct

    return productProfits, plantProductHours, plantAvailableHours


def writeDataOpenpyxl(soln, objVal) -> None:
    """
    Write the solution back to the Excel file using openpyxl.

    Parameters
    ----------
    soln : dict
        The optimal solutions
        - Keys: variable names
        - Values: optimal decision variable values
    objVal : float
        The optimal objective function value
    """
    # Load a workbook from DATA_PATH
    outputBook = load_workbook(constant.DATA_PATH)

    # Find the sheet to write the solution
    outputSheet = outputBook[constant.SHEET_NAME]

    # Write the optimal solutions to the outputSheet
    for product in constant.PRODUCT_NAMES:
        outputCol = getattr(constant, f"OUTPUT_{product.upper()}_COL")
        outputSheet.cell(constant.OUTPUT_ROW, outputCol, soln[product])
    outputSheet.cell(constant.OUTPUT_ROW, constant.OUTPUT_PROFIT_COL, objVal)

    # Save the workbook
    outputBook.save(constant.DATA_PATH)
